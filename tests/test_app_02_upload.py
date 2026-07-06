"""app_02_upload.py 비즈니스 로직 단위 테스트.

테스트 대상:
- _fetch_distinct_values: 고유 값 조회, 정렬, 오류 처리
- make_tamplate_xlsx: 템플릿 생성, ABC class 제거, 사업부/채널 드롭다운 추가
- read_upload_xl: ABC_CLASS 컬럼 누락 대응
"""

import sys
import os
import io
import pandas as pd
from unittest import mock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

sys.modules.setdefault("streamlit", mock.MagicMock())
sys.modules.setdefault("snowflake_SQL", mock.MagicMock())

import app_02_upload as app_upload

import openpyxl


class TestFetchDistinctValues:
    def test_returns_sorted_distinct_values(self):
        """고유 값을 중복 제거하고 정렬하여 반환해야 한다."""
        mock_conn = mock.MagicMock()
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_snowflake_SQL.query_to_snowflake_with_text.side_effect = None
        mock_snowflake_SQL.query_to_snowflake_with_text.return_value = pd.DataFrame(
            {"DEPT": ["B", "A", "A"]}
        )
        result = app_upload._fetch_distinct_values(
            mock_conn, "DEPT", "TESTDB.PUBLIC.MONTH_FORECAST_CONSOL"
        )
        assert result == ["A", "B"]

    def test_returns_empty_list_on_error(self):
        """조회 실패 시 빈 리스트를 반환해야 한다."""
        mock_conn = mock.MagicMock()
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_snowflake_SQL.query_to_snowflake_with_text.side_effect = Exception("DB error")
        result = app_upload._fetch_distinct_values(
            mock_conn, "DEPT", "TESTDB.PUBLIC.MONTH_FORECAST_CONSOL"
        )
        assert result == []

    def test_ignores_blank_values(self):
        """빈 문자열 및 None 값은 제외해야 한다."""
        mock_conn = mock.MagicMock()
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_snowflake_SQL.query_to_snowflake_with_text.side_effect = None
        mock_snowflake_SQL.query_to_snowflake_with_text.return_value = pd.DataFrame(
            {"DEPT": ["A", "", None, "B"]}
        )
        result = app_upload._fetch_distinct_values(
            mock_conn, "DEPT", "TESTDB.PUBLIC.MONTH_FORECAST_CONSOL"
        )
        assert result == ["A", "B"]

    def test_handles_lowercase_column_name(self):
        """Snowflake가 소문자 컬럼명을 반환해도 정상 동작해야 한다."""
        mock_conn = mock.MagicMock()
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_snowflake_SQL.query_to_snowflake_with_text.side_effect = None
        mock_snowflake_SQL.query_to_snowflake_with_text.return_value = pd.DataFrame(
            {"dept": ["X", "Y"]}
        )
        result = app_upload._fetch_distinct_values(
            mock_conn, "DEPT", "TESTDB.PUBLIC.MONTH_FORECAST_CONSOL"
        )
        assert result == ["X", "Y"]


class TestMakeTemplateXlsx:
    def _setup_snowflake_mock(self, dept_values, channel_values):
        """DEPT, CHANNEL 조회를 모의한다."""
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_engine = mock.MagicMock()
        mock_conn = mock.MagicMock()
        mock_engine.connect.return_value.__enter__.return_value = mock_conn
        mock_engine.connect.return_value.__exit__.return_value = False
        mock_snowflake_SQL.connect_snowflake.return_value = mock_engine
        mock_snowflake_SQL.query_to_snowflake_with_text.side_effect = [
            pd.DataFrame({"DEPT": dept_values}),
            pd.DataFrame({"CHANNEL": channel_values}),
        ]

    def _load_template(self):
        """make_tamplate_xlsx 결과를 openpyxl 워크북으로 반환한다."""
        xlsx_bytes = app_upload.make_tamplate_xlsx()
        return openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    def test_abc_class_column_removed(self):
        """템플릿 헤더에서 'ABC class' 컬럼이 제거되어야 한다."""
        self._setup_snowflake_mock(["A", "B"], ["CH1", "CH2"])
        wb = self._load_template()
        ws = wb["업로드 양식"]
        header = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
        assert "ABC class" not in header

    def test_dept_channel_headers_exist(self):
        """헤더에 '사업부'와 '채널' 컬럼이 존재해야 한다."""
        self._setup_snowflake_mock(["A", "B"], ["CH1", "CH2"])
        wb = self._load_template()
        ws = wb["업로드 양식"]
        header = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
        assert "사업부" in header
        assert "채널" in header

    def test_dept_channel_dropdown_values_in_hidden_sheet(self):
        """사업부/채널 드롭다운 값은 숨겨진 시트에 저장되어야 한다."""
        self._setup_snowflake_mock(["A", "B"], ["CH1", "CH2"])
        wb = self._load_template()
        hidden_ws = wb["dropdown_values"]
        assert hidden_ws.sheet_state == "hidden"
        dept_values = [hidden_ws.cell(row=i, column=1).value for i in range(1, 3)]
        channel_values = [hidden_ws.cell(row=i, column=2).value for i in range(1, 3)]
        assert dept_values == ["A", "B"]
        assert channel_values == ["CH1", "CH2"]

    def test_dept_channel_dropdowns_not_enforced(self):
        """사업부/채널 드롭다운은 강제하지 않아야 한다."""
        self._setup_snowflake_mock(["A", "B"], ["CH1", "CH2"])
        wb = self._load_template()
        ws = wb["업로드 양식"]
        non_enforced = [dv for dv in ws.data_validations.dataValidation if not dv.showErrorMessage]
        assert len(non_enforced) >= 2

    def test_status_dropdown_still_enforced(self):
        """Status 드롭다운은 여전히 강제되어야 한다."""
        self._setup_snowflake_mock(["A", "B"], ["CH1", "CH2"])
        wb = self._load_template()
        ws = wb["업로드 양식"]
        enforced = [dv for dv in ws.data_validations.dataValidation if dv.showErrorMessage]
        assert len(enforced) >= 1

    def test_generates_template_even_when_query_fails(self):
        """Snowflake 조회 실패 시에도 템플릿이 생성되어야 한다."""
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_engine = mock.MagicMock()
        mock_engine.connect.side_effect = Exception("Connection failed")
        mock_snowflake_SQL.connect_snowflake.return_value = mock_engine
        xlsx_bytes = app_upload.make_tamplate_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert "업로드 양식" in wb.sheetnames
        assert "dropdown_values" not in wb.sheetnames


class TestReadUploadXl:
    def _make_upload_excel(self):
        """업로드 양식 테스트용 엑셀 바이너리를 생성한다."""
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "업로드 양식"
        ws.append(["필수", "1: 신제품, 2: 런닝품, 3: 단종 임박", "필수", "필수", "FCST"])
        ws.append(["SKU", "Status", "사업부", "채널", "202506"])
        ws.append(["SKU001", "1", "A", "CH1", "100"])
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _setup_snowflake_mock(self):
        """PRODUCT_MASTER 조회 및 연결을 모의한다."""
        mock_snowflake_SQL = sys.modules["snowflake_SQL"]
        mock_engine = mock.MagicMock()
        mock_conn = mock.MagicMock()
        mock_engine.connect.return_value.__enter__.return_value = mock_conn
        mock_engine.connect.return_value.__exit__.return_value = False
        mock_snowflake_SQL.connect_snowflake.return_value = mock_engine
        mock_snowflake_SQL.query_to_snowflake_with_text.side_effect = None
        mock_snowflake_SQL.query_to_snowflake_with_text.return_value = pd.DataFrame({
            "품목코드": ["SKU001"],
            "요청_품목명_국문": ["Test Product"]
        })

    def test_adds_abc_class_when_missing(self):
        """템플릿에 ABC class가 없을 때 ABC_CLASS 컬럼을 None으로 추가해야 한다."""
        self._setup_snowflake_mock()
        mock_st = sys.modules["streamlit"]
        mock_st.session_state = {"user_name_kr": "테스트"}

        uploaded = self._make_upload_excel()
        result = app_upload.read_upload_xl(uploaded)

        assert result is not None
        assert "ABC_CLASS" in result.columns
        assert result["ABC_CLASS"].isna().all()
