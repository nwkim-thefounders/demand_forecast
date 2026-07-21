import streamlit as st
import pandas as pd
import datetime
import logging
import snowflake_SQL
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from dateutil.relativedelta import relativedelta
from typing import Optional

logger = logging.getLogger(__name__)

check_help_text = """
**검사 로직 설명**
1. SKU 컬럼이 비어있는 행 삭제
2. DESC 컬럼이 비어 있으면 에러 메시지 표시
3. 사업부 컬럼이 비어있으면 에러 메시지 표시
4. 채널 컬럼이 비어있으면 에러 메세지 표시
"""

read_help_text = """
**시트 읽기 로직**
1. QTY 시트를 읽어옵니다.
2. SKU 컬럼이 있는 행부터 시작합니다.
3. SKU 컬럼 이전의 컬럼은 삭제합니다.
4. "카테고리"부터 "FINAL Forecast" 라는 글자가 있는 컬럼까지 삭제합니다.
5. "FINAL Forecast" 컬럼 이후의 컬럼은 삭제합니다.
6. 첫 번째 행을 헤더로 설정합니다.
"""

col_mapping = {
    'SKU': 'SKU',
    'DESC': 'DESC',
    'Status': 'STATUS',
    'ABC class': 'ABC_CLASS',
    '사업부': 'DEPT',
    '채널': 'CHANNEL',
    'FINAL Forecast': 'FINAL_FORECAST'
}


def melt_logic(df: pd.DataFrame) -> pd.DataFrame:
    """QTY 원본 DataFrame을 wide → long 형태로 피벗(Melt)하고 DB 저장 양식으로 변환한다.

    Args:
        df (pd.DataFrame): 다수 월 컬럼을 포함한 wide 형태의 원본 DataFrame.

    Returns:
        pd.DataFrame: SIGNOFF_DT, FCST_MTH, SKU 등 DB 컬럼 구조를 갖는 long 형태 DataFrame.
    """
    df = df.copy()
    id_vars = ['카테고리', 'SKU', 'DESC', 'Status', 'ABC class', '사업부', '채널']
    df = pd.melt(
        df,
        id_vars=id_vars,
        var_name='MONTH',
        value_name="FORECAST_QTY"
    )
    df["SIGNOFF_DT"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    df["FCST_MTH"] = pd.Timestamp.now().strftime("%Y%m")
    df = df.set_index(["SIGNOFF_DT", "FCST_MTH"]).reset_index()

    df["FCST_MTH"] = pd.to_numeric(df["FCST_MTH"], errors="coerce")
    df["MONTH"] = pd.to_numeric(df["MONTH"], errors="coerce")
    df["FORECAST_QTY"] = pd.to_numeric(df["FORECAST_QTY"], errors="coerce")
    df["REGISTANT"] = st.session_state.get("user_name_kr", "")

    df = df.drop(columns="카테고리")
    df = df.rename(columns=col_mapping)
    return df


def read_origin_xl(uploaded_file: object) -> Optional[pd.DataFrame]:
    """Monthly Forecast 원본 엑셀(QTY 시트)을 파싱하여 DB 저장 형태로 반환한다.

    Args:
        uploaded_file (object): Streamlit file_uploader로 받은 업로드 파일 객체.

    Returns:
        Optional[pd.DataFrame]: 성공 시 long 형태 DataFrame, 실패 시 None.
        실패 시 st.session_state['err_msg']와 ['is_valid']에 에러 정보를 저장한다.
    """
    origin_df = pd.read_excel(uploaded_file, sheet_name="QTY")

    df = origin_df.copy()

    # 엑셀에 표기된fcst 날짜(a1셀)가 이번달과 다르면 에러 메시지 출력 처음 읽으면 첫번째 컬럼으로 인식됨
    fcst_date = pd.to_datetime(df.columns[0]).strftime("%Y-%m")
    
    if fcst_date != pd.Timestamp.now().strftime("%Y-%m"):
        st.session_state["err_msg"] = f"엑셀에 표기된 **A1셀** 날짜가 이번달({pd.Timestamp.now().strftime('%Y-%m')})이 아닙니다.\n\n날짜 업데이트 후 업로드 부탁드립니다."
        st.session_state["is_valid"] = False
        return None

    # 1. SKU 행 시작점 찾기
    sku_row_found = False
    for i, row in df.iterrows():
        if "SKU" in row.values:
            df = df.iloc[i:]
            sku_row_found = True
            break
    if not sku_row_found:
        st.session_state["err_msg"] = "시트 내에서 'SKU'가 포함된 시작 행을 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    for i, row in df.iterrows():
        if "SKU" in str(row):
            for j, col in enumerate(row):
                if "SKU" in str(col):
                    df = df.drop(df.columns[j-1], axis=1)
                    break
            break

    # 2. 카테고리 인덱스 찾기
    category_col_idx = None
    for i, row in origin_df.iterrows():
        if any("카테고리" in str(val) for val in row.values):
            for j, col in enumerate(row):
                if "카테고리" in str(col):
                    category_col_idx = j
                    break
            break
    if category_col_idx is None:
        st.session_state["err_msg"] = "시트에서 '카테고리' 컬럼 위치를 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    # 3. FINAL Forecast 인덱스 찾기
    final_forecast_col_idx = None
    sub_header_i = None
    for i, row in origin_df.iterrows():
        if any("FINAL Forecast" in str(val) for val in row.values):
            sub_header_i = i
            for j, col in enumerate(row):
                if "FINAL Forecast" == str(col).strip():
                    final_forecast_col_idx = j-1
                    break
            break

    if final_forecast_col_idx is None:
        st.session_state["err_msg"] = "시트에서 'FINAL Forecast' 컬럼 위치를 정확히 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    # 4. M-1 인덱스 찾기
    m1_col_idx = None
    for i, row in origin_df.iterrows():
        if i == sub_header_i:
            for j, col in enumerate(row):
                if j < final_forecast_col_idx + 2:
                    continue
                if isinstance(col, str):
                    m1_col_idx = j-1
                    break
            break
    if m1_col_idx is None:
        st.session_state["err_msg"] = "'FINAL Forecast' 이후에 오는 기준 데이터(M-1 등) 컬럼 위치를 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    # 데이터 자르기 공정
    try:
        # 4단계 : m1_col_idx부터 끝까지 삭제
        df = df.drop(df.columns[m1_col_idx:], axis=1)
        # 5단계 : category_col_idx부터 final_forecast_col_idx까지 삭제
        df = df.drop(df.columns[category_col_idx:final_forecast_col_idx], axis=1)

        # 첫 번째 행을 헤더로 설정
        df.columns = df.iloc[0]
        df = df[1:]
        df = df.reset_index(drop=True)
    except Exception as e:
        st.session_state["err_msg"] = f"지정된 인덱스로 컬럼을 자르는 중 오류가 발생했습니다: {e}"
        st.session_state["is_valid"] = False
        return None

    # 5. 헤더 변환 후 SKU 컬럼 존재 여부 체크
    if "SKU" not in df.columns:
        st.session_state["err_msg"] = "데이터 변환 후 'SKU' 컬럼이 헤더에 존재하지 않습니다. 엑셀 파일 형식을 확인해주세요."
        st.session_state["is_valid"] = False
        return None

    # SKU가 비어있는 행 삭제
    df = df.loc[df["SKU"].notna()]

    # desc가 비어있는 행 찾기
    if "DESC" not in df.columns:
        st.session_state["err_msg"] = "'DESC' 컬럼을 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    desc_err_df = df[df["DESC"].isna()]
    if not desc_err_df.empty:
        st.session_state["err_msg"] = "QTY 시트의 DESC 컬럼에 빈 값이 있습니다."
        st.session_state["is_valid"] = False
        return df

    # 사업부가 비어있는 행 찾기
    if "사업부" not in df.columns:
        st.session_state["err_msg"] = "'사업부' 컬럼을 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    business_unit_err_df = df[df["사업부"].isna()]
    if not business_unit_err_df.empty:
        st.session_state["err_msg"] = "QTY 시트의 사업부 컬럼에 빈 값이 있습니다."
        st.session_state["is_valid"] = False
        return df

    # 채널이 비어있는 행 찾기
    if "채널" not in df.columns:
        st.session_state["err_msg"] = "'채널' 컬럼을 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    channel_err_df = df[df["채널"].isna()]
    if not channel_err_df.empty:
        st.session_state["err_msg"] = "QTY 시트의 채널 컬럼에 빈 값이 있습니다."
        st.session_state["is_valid"] = False
        return df

    st.session_state["is_valid"] = True
    st.session_state["err_msg"] = ""

    try:
        df = melt_logic(df)
    except Exception as e:
        st.session_state["err_msg"] = f"데이터 구조 변경(Melt) 중 오류가 발생했습니다: {e}"
        st.session_state["is_valid"] = False
        return None
    df = df.replace(0, None)
    df = df.dropna(subset="FORECAST_QTY")

    # write_pandas / pd_writer -> to_parquet 대응: 숫자 컬럼 dtype 재보장
    df["MONTH"] = pd.to_numeric(df["MONTH"], errors="coerce")
    df["FCST_MTH"] = pd.to_numeric(df["FCST_MTH"], errors="coerce")
    df["FORECAST_QTY"] = pd.to_numeric(df["FORECAST_QTY"], errors="coerce")
    df["ABC_CLASS"] = pd.to_numeric(df["ABC_CLASS"], errors="coerce")

    st.session_state["df"] = df
    return df


def read_upload_xl(uploaded_file: object) -> Optional[pd.DataFrame]:
    """업로드 양식 엑셀을 파싱하고 Snowflake 품목명을 조인하여 반환한다.

    Args:
        uploaded_file (object): Streamlit file_uploader로 받은 업로드 파일 객체.

    Returns:
        Optional[pd.DataFrame]: 성공 시 조인 완료된 long 형태 DataFrame, 실패 시 None.
    """
    engine = snowflake_SQL.connect_snowflake()
    with engine.connect() as conn:
        pdf = snowflake_SQL.query_to_snowflake_with_text(
            'SELECT "품목코드", "요청_품목명_국문" FROM TESTDB.PUBLIC.PRODUCT_MASTER',
            conn=conn
        )
    df = pd.read_excel(uploaded_file, sheet_name="업로드 양식")
    header_row = None
    for i, row in df.iterrows():
        if "SKU" in row.values:
            header_row = i
            break

    if header_row is None:
        st.session_state["err_msg"] = "업로드 양식 시트에서 'SKU'가 포함된 헤더 행을 찾을 수 없습니다."
        st.session_state["is_valid"] = False
        return None

    df.columns = [str(col).replace(".0", "") for col in df.iloc[header_row]]
    df.columns = [col.upper().replace(" ", "_") for col in df.columns]
    df = df[header_row + 1:]
    if df.empty:
        st.session_state["err_msg"] = "데이터가 없습니다."
        st.session_state["is_valid"] = False
        return None

    # 데이터 프레임 양식 정리
    df = df.replace("", None)
    df = df.replace(0, None)

    if "STATUS" in df.columns:
        df["STATUS"] = df["STATUS"].fillna("").astype(str)

    # 템플릿에서 ABC class 컬럼이 제거될 수 있으므로 누락 시 None으로 추가
    if "ABC_CLASS" not in df.columns:
        df["ABC_CLASS"] = None

    # 포캐스트 월 컬럼 추출
    month_col = []
    for col in df.columns:
        try:
            pd.to_numeric(col)
        except ValueError:
            continue
        month_col.append(col)

    # 포캐스트 월에 해당하는 컬럼이 없으면 에러 발생
    if not month_col:
        st.session_state["err_msg"] = "포캐스트 월에 해당하는 컬럼을 찾을 수 없습니다.\n 예: '202606', '202607'"
        st.session_state["is_valid"] = False
        return None

    # 포캐스트 월 중 입력하지 않은 컬럼 삭제
    for col in month_col:
        month_df = df.loc[df[col].notna(), col]
        if month_df.empty:
            df = df.drop(columns=col)

    # 데이터 피벗
    df = pd.melt(
        frame=df,
        id_vars=[col for col in df.columns if col not in month_col],
        value_vars=[col for col in df.columns if col in month_col],
        value_name="FORECAST_QTY",
        var_name="MONTH",
        ignore_index=True
    )
    # FORECAST_QTY에서 None인 것 drop
    df = df.dropna(subset="FORECAST_QTY")
    # 품목명 join
    df = pd.merge(
        left=df,
        right=pdf,
        how="left",
        left_on="SKU",
        right_on="품목코드"
    )
    # 기본값 입력
    df["SIGNOFF_DT"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    df["FCST_MTH"] = datetime.datetime.now().strftime("%Y%m")
    df["REGISTANT"] = st.session_state.get("user_name_kr", "")

    # 컬럼 이름 변경 및 순서 재지정
    col_map = {
        "SIGNOFF_DT": "SIGNOFF_DT",
        "FCST_MTH": "FCST_MTH",
        "SKU": "SKU",
        "요청_품목명_국문": "DESC",
        "STATUS": "STATUS",
        "ABC_CLASS": "ABC_CLASS",
        "사업부": "DEPT",
        "채널": "CHANNEL",
        "MONTH": "MONTH",
        "FORECAST_QTY": "FORECAST_QTY",
        "REGISTANT": "REGISTANT"
    }
    df = df[[key for key in col_map.keys()]]
    df = df.rename(columns=col_map)

    # write_pandas / pd_writer -> to_parquet 대응: 숫자 컬럼 dtype 변환
    df["MONTH"] = pd.to_numeric(df["MONTH"], errors="coerce")
    df["FCST_MTH"] = pd.to_numeric(df["FCST_MTH"], errors="coerce")
    df["FORECAST_QTY"] = pd.to_numeric(df["FORECAST_QTY"], errors="coerce")
    df["ABC_CLASS"] = pd.to_numeric(df["ABC_CLASS"], errors="coerce")

    st.session_state["df"] = df
    st.session_state["is_valid"] = True
    st.session_state["err_msg"] = ""
    return df


def read_df_xlsx(uploaded_file: object) -> Optional[pd.DataFrame]:
    """업로드 엑셀의 시트명을 판별하여 적절한 파싱 함수로 분기한다.

    Args:
        uploaded_file (object): Streamlit file_uploader로 받은 업로드 파일 객체.

    Returns:
        Optional[pd.DataFrame]: 파싱 결과 DataFrame. 실패 시 None.
    """
    st.session_state["err_msg"] = ""
    xl = pd.ExcelFile(uploaded_file)
    sheet_names = xl.sheet_names
    if "업로드 양식" in sheet_names:
        return read_upload_xl(uploaded_file)
    elif "QTY" in sheet_names:
        return read_origin_xl(uploaded_file)
    else:
        st.session_state["err_msg"] = (
            "엑셀 파일의 'QTY' or '업로드 양식' 시트를 찾을 수 없습니다.\n\n"
            "**Monthly Forecast File** 엑셀 혹은 **업로드 양식** 엑셀을 업로드 해주세요"
        )
        st.session_state["is_valid"] = False
        return None


def is_sign_off(conn: object) -> bool:
    """Sign-off 토글 활성 여부 확인 후 활성화 시 기존 SIGNOFF 데이터를 삭제한다.

    Args:
        conn (object): 활성화된 SQLAlchemy Connection 객체.

    Returns:
        bool: Sign-off 토글이 활성화 상태이면 True, 아니면 False.
    """
    user = st.session_state.get("user_name_kr", "")
    month = pd.Timestamp.now().strftime("%Y%m")
    sign_off = st.session_state.get("sign_off", False)
    if sign_off:
        query = (
            f"DELETE FROM TESTDB.PUBLIC.MONTH_FORECAST_CONSOL "
            f"WHERE FCST_MTH = '{month}' AND REGISTANT = '{user}' AND SIGN_STATUS = 'SIGNOFF';"
        )
        snowflake_SQL.query_to_snowflake_with_text(query=query, conn=conn)
        return True
    return False


def _fetch_distinct_values(conn: object, column: str, table: str) -> list[str]:
    """Snowflake에서 지정 컬럼의 고유 값 목록을 조회한다.

    Args:
        conn (object): 활성화된 SQLAlchemy Connection 객체.
        column (str): 조회할 컬럼명.
        table (str): 조회할 테이블명 (스키마 포함).

    Returns:
        list[str]: 고유 값 문자열 목록. 조회 실패 시 빈 리스트 반환.
    """
    try:
        query = f'SELECT DISTINCT "{column}" FROM {table}'
        df = snowflake_SQL.query_to_snowflake_with_text(query, conn=conn)
        # Snowflake/SQLAlchemy 반환 컬럼명이 대문자/소문자로 달라질 수 있으므로
        # 첫 번째 컬럼을 직접 사용한다.
        values = df.iloc[:, 0].dropna().astype(str).unique().tolist()
        return sorted([v for v in values if v.strip()])
    except Exception as e:
        logger.error("%s DISTINCT 조회 실패: %s", column, e)
        return []


def make_tamplate_xlsx() -> bytes:
    """openpyxl로 업로드 양식 템플릿을 생성하고 바이너리 데이터로 반환한다.

    Returns:
        bytes: 생성된 엑셀 파일의 바이너리 데이터.
    """
    excel_buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "업로드 양식"

    today = datetime.datetime.now()

    # 달 추가
    month_list = [(today + relativedelta(months=i)).strftime("%Y%m") for i in range(13)]

    # DEPT, CHANNEL 드롭다운 값 조회
    dept_options = []
    channel_options = []
    try:
        engine = snowflake_SQL.connect_snowflake()
        with engine.connect() as conn:
            dept_options = _fetch_distinct_values(conn, "DEPT", "TESTDB.PUBLIC.MONTH_FORECAST_CONSOL")
            channel_options = _fetch_distinct_values(conn, "CHANNEL", "TESTDB.PUBLIC.MONTH_FORECAST_CONSOL")
    except Exception as e:
        logger.error("드롭다운 값 조회 중 오류: %s", e)

    manual = ["필수", "1: 신제품, 2: 런닝품, 3: 단종 임박", "필수", "필수"] + ["FCST" for _ in month_list]
    header = ["SKU", "Status", "사업부", "채널"] + month_list
    ws.append(manual)
    ws.append(header)

    # 서식 지정
    manual_font = Font(name="맑은 고딕", size=11, bold=False, color="FF0000")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for i in range(len(manual)):
        ws.cell(row=1, column=i+1).font = manual_font

    for i in range(len(header)):
        ws.cell(row=2, column=i+1).font = header_font
        ws.cell(row=2, column=i+1).fill = header_fill

    # Status 드롭다운 (강제)
    status_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    status_dv.showErrorMessage = True
    status_dv.errorStyle = "stop"
    status_dv.errorTitle = "입력 오류 (지정된 목록 없음)"
    status_dv.error = "드롭다운 목록에 있는 값만 입력할 수 있습니다."
    status_dv.add("B3:B1000")
    ws.add_data_validation(status_dv)

    # 사업부, 채널 드롭다운을 위한 숨겨진 시트
    if dept_options or channel_options:
        hidden_ws = wb.create_sheet("dropdown_values")
        hidden_ws.sheet_state = "hidden"
        for i, val in enumerate(dept_options, 1):
            hidden_ws.cell(row=i, column=1, value=val)
        for i, val in enumerate(channel_options, 1):
            hidden_ws.cell(row=i, column=2, value=val)

    # 사업부 드롭다운 (강제하지 않음)
    if dept_options:
        dept_dv = DataValidation(
            type="list",
            formula1=f"='dropdown_values'!$A$1:$A${len(dept_options)}",
            allow_blank=True
        )
        dept_dv.showErrorMessage = False
        dept_dv.add("C3:C1000")
        ws.add_data_validation(dept_dv)

    # 채널 드롭다운 (강제하지 않음)
    if channel_options:
        channel_dv = DataValidation(
            type="list",
            formula1=f"='dropdown_values'!$B$1:$B${len(channel_options)}",
            allow_blank=True
        )
        channel_dv.showErrorMessage = False
        channel_dv.add("D3:D1000")
        ws.add_data_validation(channel_dv)

    # 생성한 워크북을 메모리 버퍼에 저장
    wb.save(excel_buffer)

    # 버퍼의 포인터를 처음 위치로 이동
    excel_buffer.seek(0)

    # 바이너리 데이터 반환
    return excel_buffer.getvalue()


def save_btn() -> None:
    """데이터 저장 버튼 콜백. 세션의 DataFrame을 Snowflake에 INSERT한다.

    Sign-off 토글 활성 시 기존 SIGNOFF 데이터를 삭제 후 재업로드한다.
    """
    with st.spinner("데이터를 저장 중입니다..."):
        engine = snowflake_SQL.connect_snowflake()
        with engine.connect() as conn:
            df = st.session_state.get("df", None)
            if df is not None:
                df["SIGN_STATUS"] = "REGIST"

                if is_sign_off(conn=conn):
                    df["SIGN_STATUS"] = "SIGNOFF"

                # write_pandas parquet 변환 대응 최종 dtype 정리
                df["MONTH"] = pd.to_numeric(df["MONTH"], errors="coerce")
                df["FCST_MTH"] = pd.to_numeric(df["FCST_MTH"], errors="coerce")
                df["FORECAST_QTY"] = pd.to_numeric(df["FORECAST_QTY"], errors="coerce")
                df["ABC_CLASS"] = pd.to_numeric(df["ABC_CLASS"], errors="coerce")

                snowflake_SQL.input_data(conn, df, "MONTH_FORECAST_CONSOL")

        st.session_state["df"] = None
        st.session_state["is_valid"] = False
        st.session_state["err_msg"] = ""
        st.session_state["uploader_version"] = st.session_state.get("uploader_version", 0) + 1


def show_upload_page() -> None:
    """업로드 탭 페이지를 렌더링한다.

    엑셀 파일을 업로드받아 파싱 및 유효성 검사 후 Snowflake에 저장하는 UI를 제공한다.
    """
    if "err_msg" not in st.session_state:
        st.session_state["err_msg"] = ""
        st.session_state["is_valid"] = False
        st.session_state["df"] = None
        st.session_state["uploader_version"] = 0

    main_col1, main_col2, main_col3 = st.columns([0.5, 9, 0.5])
    with main_col2:
        st.title("Demand Forecasting File Upload")
        st.write(f"{st.session_state.get('user_name_kr', '')}님, 환영합니다!")
        st.write("아래 엑셀파일을 업로드 하거나 drag & drop으로 업로드 해주세요.")

        with st.container(border=False, horizontal=True, horizontal_alignment="right"):
            xlsx_data = make_tamplate_xlsx()
            st.download_button(
                label="양식 다운로드",
                data=xlsx_data,
                file_name="regist_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                key="bulk_template_download"
            )

        uploaded_file = st.file_uploader(
            label="엑셀 파일 업로드",
            type=["xlsx", "xls"],
            accept_multiple_files=False,
            key=f"upload_excel_{st.session_state.get('uploader_version', 0)}",
        )

        if uploaded_file is not None:
            try:
                df = read_df_xlsx(uploaded_file)

                # 에러 메시지가 세션에 담겨 있다면 에러박스 출력
                if st.session_state.get("err_msg"):
                    st.error(st.session_state["err_msg"])
                elif df is not None:
                    if st.session_state["df"] is not None:
                        st.caption("시트 읽어오기", help=read_help_text)
                    st.caption("검사 로직", help=check_help_text)
                    st.write("")
                    with st.container(border=False, horizontal=True):
                        with st.container(border=False, horizontal=True, horizontal_alignment="left"):
                            st.subheader("데이터 미리보기")
                        with st.container(border=False, horizontal=True, horizontal_alignment="right"):
                            if st.session_state["is_valid"]:
                                if st.session_state.get("user_role", "") == "ADMIN":
                                    st.toggle(
                                        label="Sign off",
                                        key="sign_off",
                                        help="활성화 상태로 데이터 저장시 기존 SIGN_OFF 상태의 데이터를 삭제 후 재업로드 합니다."
                                    )
                                st.button(label="데이터 저장", type="primary", on_click=save_btn)

                    st.dataframe(df, use_container_width=True)
                    st.write(len(df))

            except Exception as e:
                # 예상치 못한 시스템 치명적 에러 핸들링
                logger.error("업로드 페이지 예외 발생: %s", e)
                st.error(f"시스템 예외가 발생했습니다 (개발자 문의 필요): {e}")
        else:
            st.info("업로드된 파일이 없습니다. 업로드할 파일을 선택해주세요.")
